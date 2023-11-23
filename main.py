import pandas
import pandas as pd
print(pandas.__version__)
import openpyxl
print(openpyxl.__version__)
from sql_demo import *

from cpeslog.log_code import _logging

input_path = 'input/input.xlsx'
output_path = 'output/output.xlsx'

def read_output_file(filename):
    """
    读取输入数据：外购电力
    :param filename:文件名称
    :return:读取数据结果
    """
    outwb = openpyxl.load_workbook(filename)  # 读文件
    sheetnames = outwb.sheetnames # 获取读文件中所有的sheet，通过名字的方式
    output_excel_first_sheet = outwb[sheetnames[0]]  # 获取第一个sheet内容
    return [output_excel_first_sheet, outwb]

ce_2016 = 35  # 建筑2016年国标规定每平米碳排
buliding_Area = 120000

def remove_nan(a):
    """
    求和函数
    :param a: 需要求和的数据
    :return: 计算值
    """
    list = [a for a in a if a == a]
    return list

def tanpai():
    """
    计算系统的总碳排放量
    :return:计算结果
    """
    output_excel_sheet = read_output_file(output_path)
    output_excel_first_sheet = output_excel_sheet[0]
    outwb = output_excel_sheet[1]
    try:
        total_purchased_electricity = sum(remove_nan(i_purchased_electricity))
        total_carbon_emissions = total_purchased_electricity * 0.5839 # 总碳排放量
        print(f"total_carbon_emissions:{total_carbon_emissions}")
        output_excel_first_sheet.cell(2, 1, total_carbon_emissions)
        outwb.save(output_path)
        _logging.info('输出碳排放相关变量成功')
    except BaseException as E:
        _logging.error('输出碳排放相关变量失败,错误原因为{}'.format(E))
        raise Exception

if __name__ == '__main__':
    # 把可能发生的错误语句放在try模块里，用except来处理异常
    try:
        time_now = datetime.datetime.now().strftime("%Y_%m")
        cur_time = datetime.datetime.now() - datetime.timedelta(minutes=3)
        cur_time = cur_time.strftime("%Y-%m-%d %H:%M")

        obaddr0 = ['17376', '17393', '17410']
        t_n_obix = 'ems_104_' + time_now

        data_obix0 = read_from_ems_capture(t_n_obix, obaddr0, cur_time) # df格式数据
        print(data_obix0)
        data_obix0.to_excel(input_path)

        purchased_electricity = pd.read_excel(input_path)
        i_purchased_electricity = purchased_electricity.iloc[:, 3]

        _logging.info('电网购电数据读取成功')
    except BaseException as E:
        _logging.error('电网购电数据读取失败,错误原因为{}'.format(E))
        raise Exception

    try:
        tanpai()
        _logging.info('碳排放测算模块运行成功')
    except BaseException as E:
        _logging.error('碳排放测算模块运行失败,错误原因为{}'.format(E))
        raise Exception

